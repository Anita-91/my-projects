from django.shortcuts import render
#from django.settings import FILEPATH,ROOMS_FILEPATH,LOG_FILEPATH,PEN_FILEPATH
#import serial
from openpyxl import Workbook,load_workbook
from datetime import datetime
import shutil,os
import pandas as pd
from zoneinfo import ZoneInfo

#ser = serial.Serial('/dev/ttyUSB0', 9600, timeout=1)

filepath='C:/Users/vikra/Learningpythons/serial_data_new.xlsx'
log_filepath='C:/Users/vikra/Learningpythons/log_serial.xlsx'
room_filepath = 'C:/Users/vikra/Learningpythons/roomname.xlsx'
pen_filepath = 'C:/Users/vikra/Learningpythons/log_roombk.xlsx'

id_list=[]
name_list=[]
ref_rooms={}
d=""



def read_roomsname(room_filepath):
    wb = load_workbook(room_filepath)
    sheet = wb['Room_Name']
    m_row = sheet.max_row
    m_col = sheet.max_column
    print("Before read->max_row:{} max_col:{}".format(m_row,m_col))
    id_list=[]
    name_list=[]
    for i in range(2, m_row + 1): #(2 ,6)
        c2 = sheet.cell(row = i, column = 1).value
        id_list.append(c2)
        c3 = sheet.cell(row = i, column = 2).value
        name_list.append(c3)
    ref_rooms = {id_list[i]: name_list[i] for i in range(len(id_list))}
    print("ref_rooms read==>",ref_rooms)
    return ref_rooms
    

def rooms(request):
    
    ser_no_list=[]
    def serial_listener():
        ser_no_list=['NR4NS120'] # NR4NS126,NR4NS121 , NR4NS220,NR4NS226,NR4NS221
        return ser_no_list
           
    def create_excel(filepath):
        print(filepath)
        wb = Workbook()
        sheet = wb.active
        wb['Sheet'].title="CALL_ROOMS"
        sheet['A1']= 'SERIAL_NO' 
        sheet['B1']= 'ID_NO'
        sheet['C1']= 'SWITCH_ID'
        sheet['D1']= 'STATION_ID'
        sheet['E1']= 'FLOOR_NO'
        sheet['F1']= 'ROOM_NO'
        sheet['G1']= 'CALL_TYPE'
        sheet['H1']= 'DATE'
        sheet['I1']= 'CALL_TIME'
        sheet['J1']= 'ATTEND_TIME'
        sheet['K1']= 'RESET_TIME'
        sheet['L1']= 'CALL_ATT_TIME_DIFF'
        sheet['M1']= 'ATT_RESET_TIME_DIFF'
        sheet['N1']= 'CALL_RESET_TIME_DIFF'
        sheet['O1']= 'TIME'
        wb.save(filepath)
        print("Created excel file")
        print("---------")

    def write_excel_serialno(ser_no_list): 
        ##writing ser_no into excel  file
        
        wb = load_workbook(filepath)
        sheet = wb["CALL_ROOMS"]
        rows =[[l] for l in ser_no_list]
        for row in rows:  
            sheet.append(row)
        
        t=datetime.now().strftime("%H:%M:%S") 
        d=datetime.now().strftime("%d/%m/%Y") 
        print("date: {} time:{}".format(d, t))
        wb.save(filepath)
        return d,t
        
    def update_excel_serialno(ref_rooms,d,t):
        
        wb = load_workbook(filepath)
        sheet = wb["CALL_ROOMS"]
        m_row = sheet.max_row
        for i in range(2, m_row + 1): 
            s_no = sheet.cell(row = i, column = 1)
            serial_no=s_no.value
            print("serial_no",serial_no)
            if len(serial_no) == 8:
                sheet.cell(row = i, column = 2).value=serial_no[-3]
                print("id_no:",serial_no[-3])
                sheet.cell(row = i, column = 3).value=serial_no[-3]
                print("swid:",serial_no[-3])
                sheet.cell(row = i, column = 4).value=serial_no[-2]
                print("stid:",serial_no[-2])
            elif len(serial_no) == 9:
                if int(serial_no[-3:-1]) == 10:
                    sheet.cell(row = i, column = 2).value=serial_no[-4:-3]
                    print("id_no:",serial_no[-4:-3])
                    sheet.cell(row = i, column = 3).value=serial_no[-4:-3]
                    print("swid:",serial_no[-4:-3])
                    sheet.cell(row = i, column = 4).value=serial_no[-3:-1]
                    print("stid:",serial_no[-3:-1])
                else:
                    sheet.cell(row = i, column = 2).value=serial_no[-4:-2]
                    sheet.cell(row = i, column = 3).value=serial_no[-4:-2]                  
                    sheet.cell(row = i, column = 4).value=serial_no[-2]
            elif len(serial_no) == 10:
                sheet.cell(row = i, column = 2).value=serial_no[5:7]
                print("swid:",serial_no[5:7])
                sheet.cell(row = i, column = 3).value=serial_no[5:7]
                print("swid:",serial_no[5:7])
                sheet.cell(row = i, column = 4).value=serial_no[7:9]
                print("stid:",serial_no[7:9])
                
            room_no = sheet.cell(row = i, column = 6)
            for key,n in ref_rooms.items():
                if key==int(sheet.cell(row = i, column = 3).value):
                  room_no.value = str(n)
                  break
            
            flr=room_no.value
            sheet.cell(row = i, column = 5).value=flr[0]
            sheet.cell(row = i, column = 7).value=serial_no[-1]
            print("call_type:",serial_no[-1])
        #end
        sheet.cell(row = m_row, column = 8).value = d
        print("date",d,type(d))
        sheet.cell(row = m_row, column = 15).value = t 
        print("t:",t,type(t))
        wb.save(filepath)
        print("Update excel done......")      
    
    def update_callroom():
        
        df=pd.read_excel(filepath,sheet_name='CALL_ROOMS',usecols='F,O')
        final_dict =df.groupby('ROOM_NO')['TIME'].apply(list).to_dict()
        print("final_dict:",final_dict)
        
        wb = load_workbook(filepath)
        sheet = wb["CALL_ROOMS"]
        m_row = sheet.max_row
        m_col = sheet.max_column
        print("Before read->max_row:{} max_col:{}".format(m_row,m_col))
        
        def differ(start_time,end_time):
            t1 = datetime.strptime(start_time, "%H:%M:%S")
            t2 = datetime.strptime(end_time, "%H:%M:%S")
            delta = t2 - t1
            diff=str(delta)
            print("diff:::",diff,type(diff))
            return diff
        
        for i in range(2, m_row+1): 
            room_no = sheet.cell(row = i, column = 6).value
            call_type = sheet.cell(row = i, column = 7).value
            if call_type == '1': 
                for k,v in final_dict.items():
                    if str(k) == room_no:
                      print(final_dict[k])
                      times=final_dict[k]
                      print("c:",times[0])
                      print("a:",times[1])
                      print("r:",times[2])
                      sheet.cell(row = i, column = 9).value = times[0]
                      sheet.cell(row = i, column = 10).value = times[1]
                      sheet.cell(row = i, column = 11).value = times[2]
                      sheet.cell(row = i, column = 12).value = differ(times[0],times[1])
                      sheet.cell(row = i, column = 13).value = differ(times[1],times[2])
                      sheet.cell(row = i, column = 14).value = differ(times[0],times[2])
        wb.save(filepath)
        
    def read_callroom(): 
        
        df=pd.read_excel(filepath,sheet_name='CALL_ROOMS',usecols='A:N')
        print("df\n",df)
        new_df = df.drop(df[(df['CALL_TYPE'] == 0) | (df['CALL_TYPE'] == 6) ].index,inplace = False)
        print("droped df\n:",new_df)
        new_df.to_excel(log_filepath, sheet_name='LOG_ROOMS', index = None, header=True)
        path = 'C:/Users/vikra/Learningpythons'
        if os.path.exists(path):
           print(f"The path '{path}' exists.")
           shutil.copy(log_filepath,pen_filepath)
           print(f"Room backup is ready now:: '{pen_filepath}'")
        else:
           print(f"The path '{path}' does not exist.")
        
        
    def read_excel_serialno(): 
        wb = load_workbook(filepath)
        sheet = wb["CALL_ROOMS"]
        m_row = sheet.max_row
        m_col = sheet.max_column
        print("Before read->max_row:{} max_col:{}".format(m_row,m_col))
        
        room_dict={}
            
        for i in range(2, m_row+1): 
            room_no = sheet.cell(row = i, column = 6).value
            flag = sheet.cell(row = i, column = 7).value
            if room_no not in room_dict.keys():
                print("if-->room_no:{} ,flag:{}".format(room_no,flag))
                if flag == '0':
                    room_dict[room_no]=flag
            else:
                if flag=='6':
                    room_dict[room_no]=flag
                    
                else:
                    room_dict.pop(room_no)
        myKeys = list(room_dict.keys())
        myKeys.sort()
        rooms_dict = {i: room_dict[i] for i in myKeys}
        wb.save(filepath)   
        total_rooms=len(rooms_dict.keys())
        myvalues = list(rooms_dict.values())
        flag="".join(myvalues)
        return rooms_dict,total_rooms,flag
 
    while True: 
        print("MainLoop starts.....")
        #ser_no_list=serial_listener()
        ser_no_list=['NR4NS126']
        print("got serial_no_list from serial():::",ser_no_list)
        if ser_no_list==None: 
            print("Got None value.So continue mainloop")
            #continue
        else:
            if len(ser_no_list)>0: 
                print("serial_no_list",ser_no_list)
                
                ref_rooms=read_roomsname(room_filepath)
                if os.path.exists(filepath):
                    print("Serial EXcel file Already Exist...")
                else:
                    print("Excel file Not Exist..")
                    create_excel(filepath)
                print("------Writing data--------")
                d,t=write_excel_serialno(ser_no_list)
                print("------Updating data--------")       	
                update_excel_serialno(ref_rooms,d,t)
                update_callroom()
                read_callroom()
                rooms_dict,total_rooms,flag=read_excel_serialno()
                print("========READ DONE==========")
                print("room_dict={} no_of_rooms={} flag:{}".format(rooms_dict,total_rooms,flag))
                return render(request,'call_room.html',{'room_dict':rooms_dict,'total_rooms':total_rooms,'flag':flag})
            else:
                print("No serialno Mainpg")
                
                


