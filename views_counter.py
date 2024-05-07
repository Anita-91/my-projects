from django.shortcuts import render
import serial
import openpyxl


ser = serial.Serial('COM4', 9600, timeout=1)
file_path='C:/Users/ADMIN/LearnPython/serial_data.xlsx'
c_list=[]
t_list=[]
f_list=[]

def rooms(request):
    count=1
    ser_no_list=[]
    def serial_listener():
        if not ser.isOpen():
            ser.open()
            
        print("=======New Serial Input==================")    
        print("ser.reset_input======>",ser.reset_input_buffer())
        print("=========================")
        while True:
            if ser.in_waiting > 0:
                print("ser.in_waiting===>",ser.in_waiting)
                print("-------Reading liness---")
                ser_no_list=[]
                line = ser.readline().decode('utf-8').rstrip()
                #print("serial_no:::{} type:{}".format(line,type(line)))
                print("len(line) == ",len(line))
                if(len(line) == 0):
                    continue
                ser_no_list.append(str(line))
                #print("------------DEcode done---------")
                print("ser_no_list:::",ser_no_list)
                return ser_no_list
            # else:
            #     print("Nothing from Serial Port")
            #     break


    def write_excel_serialno(ser_no_list,count): 
        if count==1:
            wb = openpyxl.Workbook()
            wb['Sheet'].title="CALLROOM"
            wb.create_sheet(title="SHOWROOM")
            print(wb.sheetnames)
            sheets=wb.sheetnames
            for s_name in sheets:
                sheet=wb[s_name]
                sheet['A1'] ='SERIAL_NO' 
                sheet['B1']= 'COUNTER'
                sheet['C1']= 'TOKEN'
                sheet['D1']='FLAG'
                #print(s_name)
                rows =[[l] for l in ser_no_list]
                for row in rows:  
                    sheet.append(row)
            wb.save(file_path)
            print("write if---count :",count)
        else:
            wb = openpyxl.load_workbook(file_path)
            sheets=wb.sheetnames
            for s_name in sheets:
                sheet=wb[s_name]
                rows =[[l] for l in ser_no_list]
                for row in rows:  
                    sheet.append(row)
            wb.save(file_path)
            print("write else---count :",count)
            
            
            
            

    def update_excel_serialno():
        ###### read serial no and and add c,t,f values
        wb = openpyxl.load_workbook(file_path)
        sheets=wb.sheetnames
        for s_name in sheets:
            print('s_name',s_name)
            sheet=wb[s_name]
            m_row = sheet.max_row
            print('m_row:::',m_row)
            for i in range(2, m_row + 1): #(2 ,6)
                c1 = sheet.cell(row = i, column = 1)
                serial_no=c1.value
                #counter
                c2 = sheet.cell(row = i, column = 2)
                c2.value=serial_no[1:2]
                #token
                c3 = sheet.cell(row = i, column = 3)
                c3.value=serial_no[2:5]
                #flag
                c4 = sheet.cell(row = i, column = 4)
                c4.value=serial_no[-1]
        wb.save(file_path)
        print("Update excel done......")
        
     
    def read_excel_serialno():
        wb = openpyxl.load_workbook(file_path)
        sheet = wb["SHOWROOM"]
        m_row = sheet.max_row
        print("max_row::",m_row)
        m_col = sheet.max_column
        print("Before read->max_row:{} max_col:{}".format(m_row,m_col))
        c_list=[]
        t_list=[]
        f_list=[]
        if m_row == 6:
            del_row = m_row-4
            print("del-row",del_row)
            sheet.delete_rows(idx=del_row)
            print("if max_row=6::",m_row)
            for i in range(2, m_row): #(2 ,6)
                #counter
                c2 = sheet.cell(row = i, column = 2).value
                c_list.append(c2)
                #token
                c3 = sheet.cell(row = i, column = 3).value
                t_list.append(c3)
                #flag
                c4 = sheet.cell(row = i, column = 4).value
                f_list.append(c4)
            
        else:
            print("elseif max_row<6::",m_row)
            for i in range(2, m_row + 1): #(2 ,6)
                #counter
                c2 = sheet.cell(row = i, column = 2).value
                c_list.append(c2)
                #token
                c3 = sheet.cell(row = i, column = 3).value
                t_list.append(c3)
                #flag
                c4 = sheet.cell(row = i, column = 4).value
                f_list.append(c4)
        print("c_list={} t_list={} f_list={}".format(c_list,t_list,f_list))
        room = {t_list[i]: c_list[i] for i in range(len(t_list))}
        print("rooms:::",room)
        flag="".join(f_list)
        print("flags:::",flag)
        return room,flag
        print("After read->max_row:{} max_col:{}".format(m_row,m_col))
        wb.save(file_path)  

    ##### main pg

    
    print("Start count:",count)
    while True: 
        ser_no_list=serial_listener()
        if len(ser_no_list)>0: 
            print("serial_no_list",ser_no_list)
            print("------Writing data--------") 
            print("count:",count)
            write_excel_serialno(ser_no_list,count)
            print("=======WRITE DONE===========") 
            count+=1
            update_excel_serialno()
            print("------Reading data--------") 
            room,flag=read_excel_serialno()
            print("========READ DONE==========")
        else:
            print("Not found serial no")
    print("after read over::",room,flag)
    return render(request,'counter_page.html',{'room':room,'flag':flag})
    
